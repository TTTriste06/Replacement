import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def adjust_column_width(ws, max_width=70):
    """
    自动调整工作表中每列的宽度，适配内容长度。
    忽略第一行（通常用于填充非数据内容），以第二行起为基准。
    """
    for col_cells in ws.iter_cols(min_row=2):  # 从第二行开始
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 8, max_width)


def highlight_replaced_names_in_main_sheet(ws, replaced_names: list[str], name_col_header: str = "品名", header_row_idx: int = 2):
    """
    只标红主计划 sheet 中所有品名在 replaced_names 中的整行的前三列。

    参数：
        ws: openpyxl 的 worksheet 对象（主计划）
        replaced_names: 替换过的新名字列表
        name_col_header: 表头中品名字段名称，默认是“品名”
        header_row_idx: 表头所在的行号（默认第 2 行）
    """
    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

    # 获取表头行
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row_idx]]

    if name_col_header not in header_row:
        raise ValueError(f"❌ 未找到“{name_col_header}”列，无法标红替换新品名。")

    name_col_idx = header_row.index(name_col_header) + 1  # openpyxl 列从 1 开始

    # 遍历数据行（从 header 下一行开始）
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        cell_value = str(row[name_col_idx - 1].value).strip()
        if cell_value in replaced_names:
            for cell in row[:3]:  # 只标红前三列 A, B, C
                cell.fill = red_fill




