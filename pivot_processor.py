import re
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import OUTPUT_FILENAME_PREFIX
from excel_utils import adjust_column_width, highlight_replaced_names_in_main_sheet
from mapping_utils import (
    clean_mapping_headers, 
    replace_all_names_with_mapping, 
    apply_mapping_and_merge, 
    apply_extended_substitute_mapping,
    apply_all_name_replacements
)

class PivotProcessor:
    def process(self, uploaded_files: dict, output_buffer, additional_sheets: dict = None):
        """
        替换品名、新建主计划表，并直接写入 Excel 文件（含列宽调整、标题行）。
        """
        # === 标准化上传文件名 ===
        self.dataframes = {}
        for filename, file_obj in uploaded_files.items():
            matched = False
            for keyword, standard_name in FILE_KEYWORDS.items():
                if keyword in filename:
                    self.dataframes[standard_name] = pd.read_excel(file_obj)
                    matched = True
                    break
            if not matched:
                st.warning(f"⚠️ 上传文件 `{filename}` 未识别关键词，跳过")

        # === 标准化新旧料号表 ===
        self.additional_sheets = additional_sheets
        mapping_df = self.additional_sheets.get("赛卓-新旧料号")
        if mapping_df is None or mapping_df.empty:
            raise ValueError("❌ 缺少新旧料号映射表，无法进行品名替换。")

        # 创建新的 mapping_semi：仅保留“半成品”字段非空的行
        mapping_semi = mapping_df[
            ["旧晶圆品名", "旧规格", "旧品名", "新晶圆品名", "新规格", "新品名", "半成品"]
        ]
        mapping_semi = mapping_semi[~mapping_df["半成品"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        
        # 去除“品名”为空的行
        mapping_new = mapping_df[
            ["旧晶圆品名", "旧规格", "旧品名", "新晶圆品名", "新规格", "新品名"]
        ]
        mapping_new = mapping_new[~mapping_df["新品名"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        mapping_new = mapping_new[~mapping_new["旧品名"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        
        # 去除“替代品名”为空的行，并保留指定字段
        mapping_sub1 = mapping_df[
            ["新晶圆品名", "新规格", "新品名", "替代晶圆1", "替代规格1", "替代品名1"]
        ]
        mapping_sub1 = mapping_sub1[~mapping_df["替代品名1"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        mapping_sub1.columns = [
            "新晶圆品名", "新规格", "新品名", 
            "替代晶圆", "替代规格", "替代品名"
        ]


        mapping_sub2 = mapping_df[
            ["新晶圆品名", "新规格", "新品名", "替代晶圆2", "替代规格2", "替代品名2"]
        ]
        mapping_sub2 = mapping_sub2[~mapping_df["替代品名2"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        mapping_sub2.columns = [
            "新晶圆品名", "新规格", "新品名",
            "替代晶圆", "替代规格", "替代品名"
        ]

        mapping_sub3 = mapping_df[
            ["新晶圆品名", "新规格", "新品名", "替代晶圆3", "替代规格3", "替代品名3"]
        ]
        mapping_sub3 = mapping_sub3[~mapping_df["替代品名3"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        mapping_sub3.columns = [
            "新晶圆品名", "新规格", "新品名",
            "替代晶圆", "替代规格", "替代品名"
        ]
        

        mapping_sub4 = mapping_df[
            ["新晶圆品名", "新规格", "新品名", "替代晶圆4", "替代规格4", "替代品名4"]
        ]
        mapping_sub4 = mapping_sub4[~mapping_df["替代品名4"].astype(str).str.strip().replace("nan", "").eq("")].copy()
        mapping_sub4.columns = [
            "新晶圆品名", "新规格", "新品名",
            "替代晶圆", "替代规格", "替代品名"
        ]

       

        
        # === 构建主计划 ===
        headers = ["晶圆品名", "规格", "品名", "封装厂", "封装形式", "PC"]
        main_plan_df = pd.DataFrame(columns=headers)


        name_unfulfilled = []
        name_forecast = []

        if df_unfulfilled is not None and not df_unfulfilled.empty:
            col_name = FIELD_MAPPINGS["赛卓-未交订单"]["品名"]
            name_unfulfilled = df_unfulfilled[col_name].astype(str).str.strip().tolist()
            
        all_names = pd.Series(name_unfulfilled + name_forecast)
        all_names = replace_all_names_with_mapping(all_names, mapping_new, mapping_df)
        main_plan_df = main_plan_df.reindex(index=range(len(all_names)))
        if not all_names.empty:
            main_plan_df["品名"] = all_names.values

        ## == 规格和晶圆 ==
        main_plan_df = fill_spec_and_wafer_info(
            main_plan_df,
            self.dataframes,
            self.additional_sheets,
            mapping_semi, 
            FIELD_MAPPINGS
        )

        ## == 封装厂，封装形式和PC ==
        main_plan_df = fill_packaging_info(
            main_plan_df,
            dataframes=self.dataframes,
            additional_sheets=self.additional_sheets
        )

        
        ## == 替换新旧料号、替代料号 ==
        all_replaced_names = set()  # 用 set 累计替换的新品名
        df_new = self.additional_sheets["赛卓-安全库存"]
        df_new, replaced_main = apply_mapping_and_merge(df_new, mapping_new, FIELD_MAPPINGS["赛卓-安全库存"])
        df_new, replaced_sub1 = apply_extended_substitute_mapping(df_new, mapping_sub1, FIELD_MAPPINGS["赛卓-安全库存"])
        df_new, replaced_sub2 = apply_extended_substitute_mapping(df_new, mapping_sub2, FIELD_MAPPINGS["赛卓-安全库存"])
        df_new, replaced_sub3 = apply_extended_substitute_mapping(df_new, mapping_sub3, FIELD_MAPPINGS["赛卓-安全库存"])
        df_new, replaced_sub4 = apply_extended_substitute_mapping(df_new, mapping_sub4, FIELD_MAPPINGS["赛卓-安全库存"])
        self.additional_sheets["赛卓-安全库存"] = df_new
        all_replaced_names.update(replaced_main)
        all_replaced_names.update(replaced_sub1)
        all_replaced_names.update(replaced_sub2)
        all_replaced_names.update(replaced_sub3)
        all_replaced_names.update(replaced_sub4)
        
        all_replaced_names = sorted(all_replaced_names)

         
        # === 写入 Excel 文件（主计划）===
        timestamp = datetime.now().strftime("%Y%m%d")
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            main_plan_df = clean_df(main_plan_df)
            main_plan_df.to_excel(writer, sheet_name="主计划", index=False, startrow=1)
            
            #写入主计划
            ws = writer.book["主计划"]
            ws.cell(row=1, column=1, value=f"主计划生成时间：{timestamp}")

            format_monthly_grouped_headers(ws)

            adjust_column_width(ws)


            # 设置字体加粗，行高也调高一点
            bold_font = Font(bold=True)
            ws.row_dimensions[2].height = 35
    
            # 遍历这一行所有已用到的列，对单元格字体加粗、居中、垂直居中
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = bold_font
                # 垂直水平居中
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动筛选
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A2:{last_col_letter}2"
        
            # 冻结
            ws.freeze_panes = "D3"

            append_all_standardized_sheets(writer, uploaded_files, self.additional_sheets)
            
            # 透视表
            standardized_files = standardize_uploaded_keys(uploaded_files, RENAME_MAP)
            parsed_dataframes = {
                filename: pd.read_excel(file)  # 或提前 parse 完成的 DataFrame dict
                for filename, file in standardized_files.items()
            }
            pivot_tables = generate_monthly_pivots(parsed_dataframes, pivot_config)
            for sheet_name, df in pivot_tables.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
            # 写完后手动调整所有透视表 sheet 的列宽
            for sheet_name, df in pivot_tables.items():
                ws = writer.book[sheet_name]
                for col_cells in ws.columns:
                    max_length = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length * 1.2 + 10
                    ws.column_dimensions[col_letter].width = min(adjusted_width, 50)


        output_buffer.seek(0)

       
    def set_additional_data(self, sheets_dict):
        """
        设置辅助数据表，如 预测、安全库存、新旧料号 等
        """
        self.additional_sheets = sheets_dict or {}
    
        # ✅ 对新旧料号进行列名清洗
        mapping_df = self.additional_sheets.get("赛卓-新旧料号")
        if mapping_df is not None and not mapping_df.empty:
            try:
                cleaned = clean_mapping_headers(mapping_df)
                self.additional_sheets["赛卓-新旧料号"] = cleaned
            except Exception as e:
                raise ValueError(f"❌ 新旧料号表清洗失败：{e}")

